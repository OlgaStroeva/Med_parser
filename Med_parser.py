import tkinter as tk
from tkinter import filedialog
from tkinter import scrolledtext
import docx
from striprtf.striprtf import rtf_to_text
import os
import subprocess
import platform
import base64
import vertexai
from vertexai.generative_models import GenerativeModel, Part
import vertexai.preview.generative_models as generative_models

prompt1 = """
Мне требуется разбить текст на пять частей: «Диагностированные заболевания», «Проведенные исследования», «Проведенное лечение», «Рекомендации» и «Служебная информация». Учти, что абзацы, относящиеся к одной части, идут подряд и не могут быть разделены абзацем, относящимся к другой части. Это не касается «Служебной информации», она может находиться и в начале, и в конце текста.
Мне нужно будет представить информацию в табличном виде. 
В таблице 'Диагностированные заболевания' выделяй в отдельные строки только диагнозы и подробности диагнозов. В этой таблице две колонки: 'название диагноза' и 'информация о диагнозе'. В 'названии диагноза' находится только наименование поставленного диагноза. Подробное описание, локализация, степень, стадия и т.п. находятся в колонке 'информация о диагнозе'. 
К исследованиям относятся анализы, исследования и консультации с врачами (если они находятся в последовательности абзацев, отнесенных к «Проведенным исследованиям»). В таблице «Проведенные исследования» три колонки: 'исследование', 'дата исследования' и 'результат исследования'. Раздели ячейки колонки 'результат исследования' на несколько строк, т.е. каждый показатель исследования должен находиться на отдельной строке. 
В таблице 'Проведенное лечение' три колонки: 'препарат', 'дозировка' и 'прием'. 'Дозировка' и 'прием' могут отсутствовать в строке. К 'дозировке' относятся данные о дозировке принимаемого препарата. К 'приему' относятся все данные о сроках и порядке приема. Если в одной строке указано несколько дозировок - выдели для новой дозировки информацию о приеме и вынеси в отдельную строку таблицы. Название препарата дублируется.В части «Рекомендации» одна колонка, каждая рекомендация находится на отдельной строке. 
В части «Служебная информация» содержатся сведения о пациенте, медицинском учреждении и врачах. 
Текст:
"""
promt = ""

file_name = ''

import openpyxl
from openpyxl.styles import Font
import re

def parse_string(data):
    lines = data.strip().split('\n')
    parsed_data = []
    current_label = None
    headers = []
    rows = []

    for line in lines:
        if line.startswith('##'):
            if current_label or headers or rows:
                parsed_data.append((current_label, headers, rows))
                headers = []
                rows = []
            current_label = line[2:].strip()
        elif line.startswith('|'):
            cells = line.split('|')[1:-1]
            if '---' in cells:
                continue
            if not headers:
                headers = cells
            else:
                rows.append(cells)
        elif line.strip() == '':
            rows.append([])  # Append an empty row for visual separation

    parsed_data.append((current_label, headers, rows))
    return parsed_data

def apply_formatting(cell, text):
    bold_pattern = re.compile(r'\*\*(.*?)\*\*')
    bold_segments = bold_pattern.findall(text)
    text = bold_pattern.sub(r'\1', text)
    
    cell.value = text
    if bold_segments:
        cell.font = Font(bold=True)
    return cell

def write_to_excel(parsed_data, filename):
    wb = openpyxl.Workbook()
    ws = wb.active

    start_row = 1
    for label, headers, rows in parsed_data:
        if label:
            ws['A' + str(start_row)] = label
            ws['A' + str(start_row)].font = Font(italic=True)
            start_row += 2  # Leave a blank row after the label
        
        for col_num, header in enumerate(headers, start=1):
            cell = ws.cell(row=start_row, column=col_num)
            cell.value = header
        
        start_row += 1
        
        for row in rows:
            for col_num, cell_value in enumerate(row, start=1):
                cell = ws.cell(row=start_row, column=col_num)
                apply_formatting(cell, cell_value)
            start_row += 1

    wb.save(filename)

def open_file(filepath):
    if platform.system() == 'Windows':
        os.startfile(filepath)
    elif platform.system() == 'Darwin':  # macOS
        subprocess.call(['open', filepath])
    else:  # Linux
        subprocess.call(['xdg-open', filepath])


def multiturn_generate_content():
  vertexai.init(project="1092377535514", location="asia-northeast3")
  model = GenerativeModel(
    "projects/1092377535514/locations/asia-northeast3/endpoints/5257816225513209856",
  )
  response = model.generate_content(promt + '\nПредоставь таблицу "Рекомендации", "Диагностированные заболевания" и "Проведенное лечение".')
  response += model.generate_content(promt + '\nПредоставь таблицу "Проведенные исследования".')
  parsed_data = parse_string(response)
  write_to_excel(parsed_data, file_name + '.xlsx')
  open_file(file_name + '.xlsx')

generation_config = {
    "max_output_tokens": 5048,
    "temperature": 1,
    "top_p": 1,
}

safety_settings = {
    generative_models.HarmCategory.HARM_CATEGORY_HATE_SPEECH: generative_models.HarmBlockThreshold.BLOCK_MEDIUM_AND_ABOVE,
    generative_models.HarmCategory.HARM_CATEGORY_DANGEROUS_CONTENT: generative_models.HarmBlockThreshold.BLOCK_MEDIUM_AND_ABOVE,
    generative_models.HarmCategory.HARM_CATEGORY_SEXUALLY_EXPLICIT: generative_models.HarmBlockThreshold.BLOCK_MEDIUM_AND_ABOVE,
    generative_models.HarmCategory.HARM_CATEGORY_HARASSMENT: generative_models.HarmBlockThreshold.BLOCK_MEDIUM_AND_ABOVE,
}


def read_docx(file_path):
    doc = docx.Document(file_path)
    full_text = []
    for para in doc.paragraphs:
        full_text.append(para.text)
    return '\n'.join(full_text)

def read_rtf(file_path):
    with open(file_path, 'r', encoding='utf-8') as file:
        rtf_text = file.read()
    return rtf_to_text(rtf_text)

def read_doc(file_path):
    try:
        import pypandoc
        return pypandoc.convert_file(file_path, 'plain')
    except ImportError:
        return "pypandoc is not installed."

def upload_file():
    # Open a file dialog to select a file
    file_path = filedialog.askopenfilename(filetypes=[("All Files", "*.rtf;*.doc;*.docx"), 
                                                      ("RTF Files", "*.rtf"), 
                                                      ("Word Files", "*.doc;*.docx")])

    if file_path:
        file_text = ""
        _, file_extension = os.path.splitext(file_path)
        global promt, file_name
        try:
            if file_extension.lower() == ".docx":
                file_text = read_docx(file_path)
            elif file_extension.lower() == ".rtf":
                file_text = read_rtf(file_path)
            elif file_extension.lower() == ".doc":
                file_text = read_doc(file_path)
            else:
                file_text = "Unsupported file format."
        except Exception as e:
            file_text = f"An error occurred: {e}"

        # Display the file text in the text widget
        
        promt = prompt1 + file_text
       # multiturn_generate_content()
        
        text_area.delete(1.0, tk.END)
        text_area.insert(tk.END, file_text)

# Create the main window
root = tk.Tk()
root.title("Загрузить файл")



# Create a button to upload a file
upload_button = tk.Button(root, text="Загрузить файл", command=upload_file)
upload_button.pack(pady=10)

# Create a scrolled text widget to display the file text
text_area = scrolledtext.ScrolledText(root, wrap=tk.WORD, width=60, height=20)
text_area.pack(padx=10, pady=10)


# Run the application
root.mainloop()